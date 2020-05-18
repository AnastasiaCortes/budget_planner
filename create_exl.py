#!/bin/usr/env python3
import datetime
import calendar
from calendar import Calendar
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


def week_number():
''' Returns the number of the week from the date inputed'''
    dt = datetime.date(year, month, day)
    week_number = dt.isocalendar()[1]
    return week_number

def week_dates():
''' Returns dictionary key - week number, value - dates of that week'''
    weeks = []
    week_list = {}
    for mnth in range(1,13):
        for i in cal.itermonthdates(year, mnth):
            t = i.strftime("%b,%d")
            if t not in weeks:
                weeks.append(t)
    a = 0
    b = 6
    x = 1
    while b <= len(weeks):
        start = weeks[a]
        end = weeks[b]
        string = f'From: {start} - {end}'
        week_list[x] = string
        x +=1
        a += 7
        b +=7
    return week_list

def create_exl(what, amount,where):
    dict = week_dates()
    try:
        #check if file already exsists
        workbook = load_workbook(filename="budget.xlsx")
    except:
        #Creates new file if it doesnt
        workbook = Workbook()
        for key, value in dict.items():
            #Creates seporate sheet for each week of the year
            week_sheet = workbook.create_sheet(index = key, title = value)
            for week_sheet in workbook:
                fontObj1 = Font(name='Times New Roman',size= 12, bold=True)
                sheet['A1'].font = fontObj1
                week_sheet["A1"]= "Expenses"
                sheet['B1'].font = fontObj1
                week_sheet["B1"]= "Amount"
                sheet['C1'].font = fontObj1
                week_shit["C1"]= "Total"
    week = dict[week_number()]
    # Peeking a sheet according to the date was entered
    sheet = workbook[week]
    # Adding color if expenses come out of contractor salary
    if where == "y":
        sheet.append(what,amount).fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
    else:
        sheet.append(what,amount)







 def main():
     cal= calendar.Calendar()
     dt = input("When did you spend money (YYYY M D)?")
     what = input("what did you spend it on ?")
     amount = input("How much was it?")
     wr= input("Was it from contractor salary? (y/n)")
     where = wr.lower()
     date = dt.strip()
     year, month, day = (int(i) for i in date.split(' '))
     create_exl(what,amount,where)
